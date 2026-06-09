#!/usr/bin/env python3
"""Assiste l'estrazione del bilancio "sporco" e genera lo schema normalized.json.

Uso:
    python ingest.py BILANCIO.(xlsx|pdf) [--skeleton normalized.skeleton.json] [--anno 2024]

- xlsx: dumpa tutte le celle non vuote di ogni foglio (per assistere la lettura).
- pdf:  estrae il testo con pdfplumber se disponibile; altrimenti avvisa che il PDF
        (es. scansione) va letto direttamente dall'agente (vision).
Stampa anche il catalogo delle chiavi canoniche da compilare.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from fsa import mapping, model  # noqa: E402


def dump_xlsx(path: Path) -> None:
    from openpyxl import load_workbook

    # read_only: streaming senza caricare stili — regge anche export molto grandi.
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    try:
        print(f"== XLSX: {path.name} | fogli: {wb.sheetnames} ==")
        for ws in wb.worksheets:
            try:
                dims = ws.calculate_dimension()
            except ValueError:  # metadati di dimensione assenti nel foglio
                dims = "?"
            print(f"\n--- foglio '{ws.title}' (dim {dims}) ---")
            lines = [
                f"{cell.coordinate}\t{cell.value!r}"
                for row in ws.iter_rows()
                for cell in row
                if cell.value is not None
            ]
            if lines:
                print("\n".join(lines))
    finally:
        wb.close()


def dump_pdf(path: Path) -> None:
    try:
        import pdfplumber  # type: ignore
    except ImportError:
        print(
            "pdfplumber non disponibile: il PDF (probabile scansione o export) va letto "
            "direttamente dall'agente con le sue capacita' multimodali.\n"
            "Per estrazione testuale automatica: pip install pdfplumber",
            file=sys.stderr,
        )
        return
    print(f"== PDF: {path.name} ==")
    with pdfplumber.open(str(path)) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            print(f"\n--- pagina {i} ---")
            text = page.extract_text() or ""
            print(text)
            for t_index, table in enumerate(page.extract_tables() or [], start=1):
                print(f"\n[tabella {i}.{t_index}]")
                for tr in table:
                    print("\t".join("" if c is None else str(c) for c in tr))


def main() -> int:
    ap = argparse.ArgumentParser(description="Ingest assistito del bilancio sporco.")
    ap.add_argument("input", type=Path, help="file di bilancio (.xlsx o .pdf)")
    ap.add_argument("--skeleton", type=Path, help="scrive lo schema normalized.json vuoto")
    ap.add_argument("--anno", type=int, default=None, help="ultimo esercizio (Input!B2)")
    args = ap.parse_args()

    if not args.input.exists():
        print(f"File non trovato: {args.input}", file=sys.stderr)
        return 1

    suffix = args.input.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xls"):
        dump_xlsx(args.input)
    elif suffix == ".pdf":
        dump_pdf(args.input)
    else:
        print(f"Estensione non gestita: {suffix} (usa .xlsx o .pdf)", file=sys.stderr)
        return 1

    print("\n== CHIAVI CANONICHE DA COMPILARE (normalized.json) ==")
    print("Conto Economico:", list(mapping.CE_KEYS))
    print("Stato Patrimoniale:", list(mapping.SP_KEYS))
    print("Meta:", list(mapping.META_CELLS))

    if args.skeleton:
        args.skeleton.write_text(
            json.dumps(model.skeleton(args.anno), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"\nSchema scritto in: {args.skeleton}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
