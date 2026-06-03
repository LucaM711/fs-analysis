"""Lettura dei valori derivati dal foglio di calcolo gia' ricalcolato.

Il modello e' l'unica fonte di verita' per indici, grade e fido: qui NON si
ricalcola nulla, si legge ``data_only=True`` il workbook prodotto da
:func:`fsa.soffice.recalc` e si costruisce un dizionario per la narrativa e per i
rating del Report.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from . import mapping

# value (1..5) -> etichetta rating (coerente con Report!B14 e A11/A14/A17)
RATING_BY_SCORE = {
    1: "Eccellente",
    2: "Positiva",
    3: "Adeguata",
    4: "Critica",
    5: "Molto Critica",
}


def rating_label(score: Any) -> str | None:
    """Mappa un punteggio 1..5 sull'etichetta del Report. ``None`` se non calcolabile."""
    try:
        s = int(round(float(score)))
    except (TypeError, ValueError):
        return None
    s = max(1, min(5, s))
    return RATING_BY_SCORE[s]


def _v(ws, coord: str) -> Any:
    return ws[coord].value


def _pair(ws, b_coord: str, c_coord: str) -> dict[str, Any]:
    return {"y0": _v(ws, b_coord), "y1": _v(ws, c_coord)}


def read_back(recalced_xlsx: str | Path) -> dict[str, Any]:
    """Estrae il set di metriche dal workbook ricalcolato."""
    wb = load_workbook(filename=str(recalced_xlsx), data_only=True)
    rep = wb[mapping.REPORT_SHEET]
    inp = wb[mapping.INPUT_SHEET]

    economico = {
        "valore_produzione": _pair(rep, "B98", "C98"),
        "valore_produzione_var": _v(rep, "D98"),
        "ricavi_di_vendita": _pair(rep, "B99", "C99"),
        "ebit": _pair(rep, "B115", "C115"),
        "ebit_var": _v(rep, "D115"),
        "ebit_pct": _pair(rep, "B116", "C116"),
        "ebitda": _pair(rep, "B136", "C136"),
        "utile": _pair(rep, "B119", "C119"),
        "utile_var": _v(rep, "D119"),
        "ros": _pair(rep, "B133", "C133"),
        "roe": _pair(rep, "B134", "C134"),
        "roi": _pair(rep, "B135", "C135"),
        "costi_produzione": _pair(rep, "B100", "C100"),
    }
    finanziario = {
        "test_acido": _pair(rep, "B126", "C126"),
        "current_ratio": _pair(rep, "B127", "C127"),
        "rapporto_indebitamento": _pair(rep, "B129", "C129"),
        "pfn": _pair(rep, "B130", "C130"),
        "pfn_ebitda": _pair(rep, "B131", "C131"),
        "dso": _pair(rep, "B139", "C139"),
        "dpo": _pair(rep, "B140", "C140"),
        "dio": _pair(rep, "B141", "C141"),
        "ccc": _pair(rep, "B142", "C142"),
    }
    patrimoniale = {
        "totale_attivo": _pair(rep, "B73", "C73"),
        "patrimonio_netto": _pair(rep, "B76", "C76"),
        "patrimonio_netto_pct": _pair(rep, "B77", "C77"),
        "debiti_finanziari": _pair(rep, "B86", "C86"),
        "debiti_finanziari_pct": _pair(rep, "B89", "C89"),
        "disponibilita_liquide": _pair(rep, "B69", "C69"),
        "disponibilita_liquide_pct": _pair(rep, "B70", "C70"),
        "immobilizzazioni": _pair(rep, "B61", "C61"),
        "dipendenti": _pair(rep, "B120", "C120"),
    }

    sub_redd = _v(inp, "C127")
    sub_fin = _v(inp, "C135")
    sub_patr = _v(inp, "C140")
    grade_fido = {
        "grade_pesato": _v(inp, "E75"),
        "grade_finale": _v(inp, "C83"),
        "fido_massimo": _v(inp, "D83"),
        "fido_label": _v(inp, "D85"),
        "dimensione_impresa": _v(inp, "I92"),
        "sub_grade_reddituale": sub_redd,
        "sub_grade_finanziaria": sub_fin,
        "sub_grade_patrimoniale": sub_patr,
    }

    ratings = {
        # B11 (Reddituale) e B17 (Patrimoniale) sono valori da scrivere;
        # B14 (Finanziaria) e' gia' una formula nel modello.
        "reddituale": rating_label(sub_redd),
        "finanziaria": _v(rep, "B14"),
        "patrimoniale": rating_label(sub_patr),
    }

    meta = {
        "nome_azienda": _v(rep, "B1"),
        "anno": _v(inp, "B2"),
        "partita_iva": _v(rep, "B4"),
        "ateco": _v(rep, "B5"),
        "forma_giuridica": _v(rep, "B6"),
        "sede": _v(rep, "B7"),
    }

    return {
        "meta": meta,
        "economico": economico,
        "finanziario_patrimoniale": {**finanziario, **patrimoniale},
        "grade_fido": grade_fido,
        "ratings": ratings,
    }
