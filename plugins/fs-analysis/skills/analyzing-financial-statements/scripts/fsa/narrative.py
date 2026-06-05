"""Iniezione delle 4 analisi testuali e dei rating nelle celle del Report.

Sono le UNICHE celle del foglio ``Report`` che vengono scritte; il layout (celle
merge, font Syne, stili, area di stampa) resta invariato.

    A34:B43  -> Ambito Economico
    C34:D43  -> Ambito Finanziario e Patrimoniale
    A46:B55  -> Sintesi
    C46:D55  -> Grade e Fido
    B11      -> rating Reddituale   (valore)
    B17      -> rating Patrimoniale (valore)
    B14      -> rating Finanziaria  (e' una formula del modello: NON si tocca)
"""

from __future__ import annotations

import math
from dataclasses import dataclass
from typing import TYPE_CHECKING, Any

from . import mapping

if TYPE_CHECKING:  # pragma: no cover
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

# Celle (top-left della merge) per ciascun blocco testuale.
TEXT_CELLS: dict[str, str] = {
    "economico": "A34",
    "finanziario_patrimoniale": "C34",
    "sintesi": "A46",
    "grade_fido": "C46",
}

# Rating scritti come valore.
RATING_CELLS: dict[str, str] = {
    "reddituale": "B11",
    "finanziaria": "B14",
    "patrimoniale": "B17",
}

# Stime tipografiche (Syne ~12pt): altezza riga e larghezza media carattere.
_LINE_HEIGHT_PT = 15.6  # 12pt * 1.3
_DEFAULT_ROW_HEIGHT_PT = 15.0
_DEFAULT_COL_WIDTH_CHARS = 8.43
_CHAR_WIDTH_FACTOR = 1.05  # caratteri proporzionali: leggera tolleranza


@dataclass
class Capacity:
    rows: int
    cols: int
    width_chars: float
    height_pt: float
    lines: int
    chars_per_line: int
    max_chars: int


def _merged_range_of(ws: "Worksheet", anchor: str):
    for rng in ws.merged_cells.ranges:
        if rng.min_row == ws[anchor].row and rng.min_col == ws[anchor].column:
            return rng
    return None


def capacity(ws: "Worksheet", anchor: str) -> Capacity:
    """Stima la capienza (caratteri) della cella merge che parte da ``anchor``."""
    rng = _merged_range_of(ws, anchor)
    if rng is None:
        # cella singola
        r = ws[anchor].row
        c = ws[anchor].column
        min_row, max_row, min_col, max_col = r, r, c, c
    else:
        min_row, max_row, min_col, max_col = rng.min_row, rng.max_row, rng.min_col, rng.max_col

    from openpyxl.utils import get_column_letter

    width_chars = 0.0
    for col in range(min_col, max_col + 1):
        dim = ws.column_dimensions.get(get_column_letter(col))
        width_chars += dim.width if dim and dim.width else _DEFAULT_COL_WIDTH_CHARS

    height_pt = 0.0
    for row in range(min_row, max_row + 1):
        dim = ws.row_dimensions.get(row)
        height_pt += dim.height if dim and dim.height else _DEFAULT_ROW_HEIGHT_PT

    lines = max(1, int(height_pt // _LINE_HEIGHT_PT))
    chars_per_line = max(1, int(width_chars * _CHAR_WIDTH_FACTOR))
    return Capacity(
        rows=max_row - min_row + 1,
        cols=max_col - min_col + 1,
        width_chars=round(width_chars, 1),
        height_pt=round(height_pt, 1),
        lines=lines,
        chars_per_line=chars_per_line,
        max_chars=lines * chars_per_line,
    )


def estimate_lines(text: str, chars_per_line: int) -> int:
    total = 0
    for paragraph in text.split("\n"):
        total += max(1, math.ceil(len(paragraph) / chars_per_line))
    return total


def budgets(wb: "Workbook") -> dict[str, dict[str, int]]:
    """Budget di caratteri/righe per ciascun blocco, da comunicare all'agente."""
    ws = wb[mapping.REPORT_SHEET]
    out: dict[str, dict[str, int]] = {}
    for block, anchor in TEXT_CELLS.items():
        cap = capacity(ws, anchor)
        out[block] = {"max_chars": cap.max_chars, "lines": cap.lines, "chars_per_line": cap.chars_per_line}
    return out


def inject(
    wb: "Workbook",
    texts: dict[str, str] | None = None,
    ratings: dict[str, Any] | None = None,
) -> list[str]:
    """Scrive testi e rating nel Report (in place). Ritorna gli avvisi di capienza."""
    ws = wb[mapping.REPORT_SHEET]
    warnings: list[str] = []

    for block, text in (texts or {}).items():
        if block not in TEXT_CELLS:
            raise KeyError(f"Blocco testuale sconosciuto: {block!r}. Ammessi: {list(TEXT_CELLS)}")
        if text is None:
            continue
        anchor = TEXT_CELLS[block]
        ws[anchor] = str(text)
        cap = capacity(ws, anchor)
        used = estimate_lines(str(text), cap.chars_per_line)
        if used > cap.lines:
            warnings.append(
                f"'{block}' ({anchor}): testo stimato {used} righe > {cap.lines} disponibili "
                f"(~{len(str(text))}/{cap.max_chars} caratteri). Rischio troncamento nel PDF."
            )

    for which, value in (ratings or {}).items():
        if which not in RATING_CELLS:
            raise KeyError(f"Rating sconosciuto: {which!r}. Ammessi: {list(RATING_CELLS)}")
        if value is None:
            continue
        ws[RATING_CELLS[which]] = value

    wb.calculation.fullCalcOnLoad = True
    return warnings
