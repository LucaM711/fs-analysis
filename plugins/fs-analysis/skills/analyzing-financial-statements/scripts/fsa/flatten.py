"""Crea la copia di rendering: solo il foglio ``Report`` con valori statici.

Per garantire che il PDF mostri sempre gli stessi numeri indipendentemente dal
comportamento di ricalcolo del renderer, si parte dal workbook GIA' ricalcolato e
si sostituiscono le formule con i loro valori in cache, lasciando intatti stili,
area di stampa e scala. Si tiene solo il foglio ``Report`` cosi' il PDF contiene
unicamente il report (l'Input non viene esportato).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from . import mapping


def _is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def flatten_report(recalced_xlsx: str | Path, out_path: str | Path) -> Path:
    """Genera ``out_path``: solo Report, formule -> valori. Ritorna il path."""
    recalced_xlsx = Path(recalced_xlsx)
    wbf = load_workbook(filename=str(recalced_xlsx), data_only=False)
    wbv = load_workbook(filename=str(recalced_xlsx), data_only=True)

    rep_f = wbf[mapping.REPORT_SHEET]
    rep_v = wbv[mapping.REPORT_SHEET]

    for row in rep_f.iter_rows():
        for cell in row:
            if _is_formula(cell.value):
                cell.value = rep_v[cell.coordinate].value

    # Tieni solo il foglio Report.
    for name in list(wbf.sheetnames):
        if name != mapping.REPORT_SHEET:
            del wbf[name]

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wbf.save(str(out))
    return out
