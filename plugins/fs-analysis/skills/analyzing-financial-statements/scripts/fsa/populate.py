"""Scrittura dei valori canonici nel foglio ``Input`` del modello.

Scrive ESCLUSIVAMENTE le celle di input elencate in :mod:`fsa.mapping`. Non tocca
mai formule, stili o il foglio ``Report``: il layout dell'output non cambia mai.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook

from . import mapping

if TYPE_CHECKING:  # pragma: no cover
    from openpyxl.workbook.workbook import Workbook

    from .model import CanonicalBalance


def open_template(template_path: str | Path) -> "Workbook":
    """Apre il modello mantenendo le formule (``data_only=False``)."""
    return load_workbook(filename=str(template_path), data_only=False)


def populate_input(wb: "Workbook", balance: "CanonicalBalance") -> None:
    """Popola il foglio Input con i valori di ``balance`` (in place)."""
    ws = wb[mapping.INPUT_SHEET]

    # Metadati.
    meta = balance.meta
    for key, cell in mapping.META_CELLS.items():
        value = meta.get(key)
        if value is None:
            continue
        ws[cell] = value

    # Conto Economico e Stato Patrimoniale, colonna per anno.
    for year in balance.years:
        col = mapping.YEAR_COLUMNS[year]
        for key, row in mapping.CONTO_ECONOMICO_ROWS.items():
            ws[f"{col}{row}"] = balance.ce(key, year)
        for key, row in mapping.STATO_PATRIMONIALE_ROWS.items():
            ws[f"{col}{row}"] = balance.sp(key, year)

    # Forza il ricalcolo di tutte le formule all'apertura (Excel/LibreOffice).
    wb.calculation.fullCalcOnLoad = True


def build_populated(
    template_path: str | Path,
    balance: "CanonicalBalance",
    out_path: str | Path,
) -> Path:
    """Apre il modello, popola l'Input e salva in ``out_path``."""
    wb = open_template(template_path)
    populate_input(wb, balance)
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out
