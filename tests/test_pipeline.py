"""Integrazione end-to-end: richiede LibreOffice (saltato se assente)."""

from __future__ import annotations

import json

import pytest
from fsa import flatten, metrics, narrative, soffice
from fsa.model import CanonicalBalance
from fsa.populate import build_populated

pytestmark = pytest.mark.skipif(
    soffice.find_soffice() is None and not soffice.have_excel(),
    reason="LibreOffice/Excel non disponibile per il rendering",
)


def test_recalc_metrics_and_pdf(tmp_path, template_path, sample_data):
    bal = CanonicalBalance.from_dict(sample_data)
    xlsx = build_populated(template_path, bal, tmp_path / "report.xlsx")

    # ricalcolo + lettura metriche dal modello
    soffice.recalc(xlsx)
    m = metrics.read_back(xlsx)
    assert m["grade_fido"]["dimensione_impresa"] == "Piccola impresa"
    assert m["grade_fido"]["grade_finale"] == 5
    assert m["grade_fido"]["fido_massimo"] == 70000
    # EBITDA = EBIT + Ammortamenti = 200000 + 120000
    assert m["economico"]["ebitda"]["y0"] == 320000

    # iniezione testi + export PDF (solo Report, valori statici)
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, data_only=False)
    narrative.inject(
        wb,
        texts={k: f"Analisi {k}." for k in narrative.TEXT_CELLS},
        ratings={"reddituale": m["ratings"]["reddituale"], "patrimoniale": m["ratings"]["patrimoniale"]},
    )
    wb.save(xlsx)
    soffice.recalc(xlsx)
    render = flatten.flatten_report(xlsx, tmp_path / "render.xlsx")

    # il render contiene solo il foglio Report con valori statici (niente formule)
    rwb = load_workbook(render, data_only=False)
    assert rwb.sheetnames == ["Report"]
    assert not str(rwb["Report"]["B98"].value).startswith("=")

    pdf = soffice.to_pdf(render, tmp_path / "report.pdf")
    assert pdf.exists() and pdf.stat().st_size > 1000


def test_skeleton_roundtrip(sample_data):
    from fsa import model

    skel = model.skeleton(2024)
    assert set(skel["conto_economico"]) == set(sample_data["conto_economico"])
    assert set(skel["stato_patrimoniale"]) == set(sample_data["stato_patrimoniale"])
    # lo schema vuoto, serializzato, e' JSON valido
    json.loads(json.dumps(skel))
