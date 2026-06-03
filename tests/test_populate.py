"""Il popolamento scrive solo le celle di input; layout/formule restano intatti."""

from __future__ import annotations

from openpyxl import load_workbook

from fsa import mapping
from fsa.model import CanonicalBalance
from fsa.populate import build_populated


def test_populate_writes_inputs_only(tmp_path, template_path, sample_data):
    bal = CanonicalBalance.from_dict(sample_data)
    out = build_populated(template_path, bal, tmp_path / "out.xlsx")

    wb = load_workbook(out, data_only=False)
    inp = wb[mapping.INPUT_SHEET]
    rep = wb[mapping.REPORT_SHEET]

    # valori di input scritti
    assert inp["B7"].value == 2000000  # Valore della produzione y0
    assert inp["B33"].value == 1160000  # Totale attivo y0
    assert inp["C17"].value == 150000  # EBIT y1
    assert inp["B1"].value == "Acme S.r.l."  # ragione sociale

    # le formule di verifica e del Report NON sono state toccate
    assert str(inp["B52"].value).startswith("=")  # Verifica Attivo
    assert str(inp["B53"].value).startswith("=")  # Verifica Passivo
    assert rep["A33"].value.strip() == "Ambito Economico"
    assert str(rep["B98"].value).startswith("=Input!")

    # ricalcolo richiesto all'apertura
    assert wb.calculation.fullCalcOnLoad is True


def test_template_has_two_sheets(template_path):
    wb = load_workbook(template_path, data_only=False)
    assert wb.sheetnames == ["Input", "Report"]
